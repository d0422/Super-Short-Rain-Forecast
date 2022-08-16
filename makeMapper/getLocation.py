# 파이썬으로 엑셀 가공
import openpyxl
import json
wb=openpyxl.load_workbook(r'H:\vscode project\SuperShortWeatherForecast\makeMapper\data.xlsx')
print(wb)
시트=wb['1']
l=dict()
for x in range(2,3790):
    si=str(시트.cell(row=x, column=3).value)
    gu=str(시트.cell(row=x,column=4).value)
    if gu!="":
        if not l.get(si):
            l[si]=[gu]
        else:
            if gu not in l[si]:
                l[si].append(gu)
with open('./locationData.json','w', encoding='utf-8') as f:
    json.dump(l,f, ensure_ascii=False)